# -*- coding: utf-8 -*-
"""
stock_panel.py — 三仓库库存管理面板

功能：
  - 三个 Tab：试剂库 / 蛋白库 / 细胞库
  - 各仓库独立表格列
  - 模糊检索框
  - Excel 导入按钮
  - 领用 / 入库 / 编辑 / 详情 / 删除 / 冻结 / 位置
"""

from datetime import datetime

from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIcon, QColor
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLineEdit, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QInputDialog,
    QMessageBox, QFileDialog, QFormLayout, QSpinBox, QComboBox,
    QTabWidget, QWidget, QScrollArea, QDoubleSpinBox,
    QLabel as QtLabel,
)

from desktop_pet.config import ICON_PATH, LOCATION_MAP
from desktop_pet.reagents.manager import (
    ReagentManager, WAREHOUSE_REAGENT, WAREHOUSE_PROTEIN, WAREHOUSE_CELL,
    WAREHOUSE_LABELS, parse_volume_str,
)
from desktop_pet.reagents.pdf_import import PDFImportDialog


# ── Excel 列头 → 内部字段 智能映射表 ───────────────
_HEADER_ALIASES = {
    "name":              ["试剂名称", "名称", "name", "品名", "试剂", "蛋白名"],
    "batch":             ["批次", "批号", "batch", "批次号", "lot"],
    "expire":            ["效期", "有效期", "expire", "过期日期", "到期日", "expiry"],
    "stock":             ["库存", "数量", "stock", "库存量", "剩余", "qty"],
    "threshold":         ["阈值", "低库存阈值", "threshold", "报警阈值", "警戒"],
    "price":             ["单价", "价格", "price", "单价(¥)", "金额"],
    "putaway_date":      ["入库日期", "入库时间", "putaway_date", "入库"],
    "location":          ["位置", "存放位置", "location", "储位", "仓位"],
    "record_book_no":    ["记录本编号", "记录本号", "record_book_no", "记录号"],
    "project_no":        ["项目号", "项目编号", "project_no", "项目代码"],
    "volume":            ["目标体积", "volume", "容量", "加身体积", "表达体积",
                          "体积(表达体积)"],
    "molecular_weight":  ["分子量", "molecular_weight", "mw", "分子质量", "分子量(kd)",
                          "分子量(kD)"],
    "isoelectric_point": ["等电点", "isoelectric_point", "pi", "pI"],
    "extinction_coeff":  ["消光系数", "extinction_coeff", "消光", "ext coefficient"],
    "concentration":     ["蛋白浓度", "浓度", "concentration", "conc", "mg/mL",
                          "mg/ml",
                          "蛋白浓度\n(mg/mL)", "蛋白浓度(mg/mL)",
                          "蛋白浓度\n(mg/ml)", "蛋白浓度(mg/ml)",
                          "浓度(mg/mL)", "浓度(mg/ml)",
                          "浓度mg/mL", "浓度mg/ml"],
    "sample_volume":     ["取样体积", "sample_volume", "实际体积", "体积(ml)",
                          "体积(mL)", "体积mL", "体积ml",
                          "蛋白体积", "样品体积"],
    "total_amount":      ["总量", "total_amount", "总含量", "总蛋白量", "总量(mg)"],
    "titer":             ["titer", "滴度", "titer mg/l", "表达量", "titer mg/L"],
    "purity":            ["纯度", "purity", "纯度(%)"],
    "buffer_type":       ["buffer", "缓冲液", "buffer_type", "缓冲体系"],
    "shipping_info":     ["shipping1", "发货信息", "shipping", "shipping_info", "发货"],
    "specification":     ["规格", "specification", "规格型号", "spec"],
    "catalog_no":        ["货号", "catalog_no", "catalog", "产品货号", "item_no", "型号"],
    "manufacturer":      ["厂家", "manufacturer", "生产商", "brand", "品牌", "供应商"],
    "notes":             ["备注", "notes", "note", "说明", "remark", "备注信息"],
}


def _map_header(raw_header):
    """将 Excel 列头名映射为内部字段名"""
    h = raw_header.strip().lower().replace("\n", " ")
    for field, aliases in _HEADER_ALIASES.items():
        if h in [a.lower().replace("\n", " ") for a in aliases]:
            return field
    return None


def _strip_unit(val):
    """
    去除数值中的单位后缀，返回纯数值字符串。
    例：'0.824 mg/mL' → '0.824'
         '7.26+0.2 mL' → '7.26+0.2'
         '5*3' → '5*3'（无单位，原样返回）
         0.824 (float) → '0.824'
    """
    if val is None:
        return ""
    # 如果已经是数值类型，直接转字符串
    if isinstance(val, (int, float)):
        # 避免科学记数法问题
        if isinstance(val, float):
            # 保留合理精度
            s = f"{val:.6g}"
            return s
        return str(val)
    s = str(val).strip()
    if not s:
        return ""
    # 去除常见单位后缀：mg/mL, mg/ml, mL, ml, mg, μg, μL, ug, uL, g, L 等
    import re
    s = re.sub(r'\s*(mg/?m[lL]|m[gL]|m[lL]|[μu][gL]|μg|uL|[gL])\s*$', '', s, flags=re.IGNORECASE)
    return s.strip()


# ── 各仓库表格列定义 ──────────────────────────────────
# 格式：(字段名, 列标题, 宽度比例)
REAGENT_COLS = [
    ("name",          "名称",        3),
    ("batch",         "批号",        2),
    ("expire",        "效期",        2),
    ("stock",         "库存",        1),
    ("sample_volume", "体积(mL)",    2),
    ("stock_mode",    "模式",        1),
    ("threshold",     "阈值",        1),
    ("price",         "单价(¥)",     1),
    ("specification", "规格",        2),
    ("catalog_no",    "货号",        2),
    ("manufacturer",  "厂家",        2),
    ("notes",         "备注",        2),
    ("location",      "位置",        2),
    ("status",        "状态",        1),
]

PROTEIN_COLS = [
    ("name",             "蛋白名",      3),
    ("project_no",       "项目号",      1.5),
    ("batch",            "批号",        1.5),
    ("concentration",    "浓度(mg/mL)", 1.5),
    ("sample_volume",    "体积(mL)",    2),
    ("total_amount",     "总量(mg)",    1.5),
    ("purity",          "纯度",        1),
    ("buffer_type",      "Buffer",      1),
    ("expire",          "效期",        1.5),
    ("location",        "位置",        1.5),
    ("status",          "状态",        1),
]

CELL_COLS = [
    ("name",          "名称",         3),
    ("cell_line",     "细胞系",        2),
    ("passage_no",    "代次",          1),
    ("culture_medium","培养基",        2),
    ("specification", "规格",          2),
    ("catalog_no",    "货号",          2),
    ("manufacturer",  "厂家",          2),
    ("notes",         "备注",          2),
    ("stock",         "库存(管)",      1),
    ("viability",     "存活率",        1),
    ("mycoplasma",    "支原体",        1),
    ("expire",        "效期",         1.5),
    ("location",      "位置",         1.5),
    ("status",        "状态",           1),
]


class StockPanel(QDialog):
    """三仓库库存面板"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("鹰谷库存 · Silver博士")
        self.setWindowIcon(QIcon(ICON_PATH))
        self.setMinimumSize(960, 520)
        self.setAcceptDrops(True)
        self._init_ui()
        self._refresh_all()

    # ── UI 构建 ──────────────────────────────────────
    def _init_ui(self):
        layout = QVBoxLayout(self)

        # --- 检索栏 ---
        search_row = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("🔍 输入名称/批次/位置/项目号检索...")
        self.search_input.setStyleSheet(
            "font-family:'Microsoft YaHei';font-size:13px;padding:5px 8px;"
            "border:1px solid #ccc;border-radius:6px;"
        )
        self.search_input.textChanged.connect(self._on_search)
        search_row.addWidget(self.search_input)
        layout.addLayout(search_row)

        # --- Tab 页 ---
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet(
            "QTabBar::tab{font-family:'Microsoft YaHei';font-size:13px;"
            "padding:6px 18px;min-width:100px;}"
        )

        # 试剂库 Tab
        self.reagent_tab = self._create_tab(
            REAGENT_COLS, WAREHOUSE_REAGENT,
            [("📦 领用",       self._use_reagent),
             ("➕ 入库",       self._add_reagent),
             ("✏️ 编辑",      self._edit),
             ("📋 详情",      self._detail),
             ("🗑️ 删除",      self._delete),
             ("🔒 冻结/解冻", self._freeze),
             ("📍 修改位置",   self._edit_location),
             ("📥 导入Excel", self._import_excel),
             ("📄 导入PDF",   self._import_pdf),
             ("🔄 刷新",      self._refresh_all)])
        self.tabs.addTab(self.reagent_tab["widget"], "🧪 试剂库")

        # 蛋白库 Tab
        self.protein_tab = self._create_tab(
            PROTEIN_COLS, WAREHOUSE_PROTEIN,
            [("📦 领用(按体积)", self._use_protein),
             ("➕ 入库",        self._add_protein),
             ("✏️ 编辑",       self._edit),
             ("📋 详情",       self._detail),
             ("🗑️ 删除",       self._delete),
             ("🔒 冻结/解冻",  self._freeze),
             ("📍 修改位置",    self._edit_location),
             ("📥 导入Excel",  self._import_excel),
             ("📄 导入PDF",    self._import_pdf),
             ("🔄 刷新",       self._refresh_all)])
        self.tabs.addTab(self.protein_tab["widget"], "🧬 蛋白库")

        # 细胞库 Tab
        self.cell_tab = self._create_tab(
            CELL_COLS, WAREHOUSE_CELL,
            [("📦 领用",       self._use_cell),
             ("➕ 入库",       self._add_cell),
             ("✏️ 编辑",      self._edit),
             ("📋 详情",      self._detail),
             ("🗑️ 删除",      self._delete),
             ("🔒 冻结/解冻", self._freeze),
             ("📍 修改位置",   self._edit_location),
             ("📥 导入Excel", self._import_excel),
             ("📄 导入PDF",   self._import_pdf),
             ("🔄 刷新",      self._refresh_all)])
        self.tabs.addTab(self.cell_tab["widget"], "🧫 细胞库")

        layout.addWidget(self.tabs)

    def _create_tab(self, col_defs, wh_type, buttons_cfg):
        """创建一个仓库 Tab 页，返回 {widget, table, wh_type}"""
        tab_widget = QWidget()
        tab_layout = QVBoxLayout(tab_widget)

        # 表格
        table = QTableWidget()
        table.setColumnCount(len(col_defs))
        table.setHorizontalHeaderLabels([c[1] for c in col_defs])
        header = table.horizontalHeader()
        for i, c in enumerate(col_defs):
            header.setSectionResizeMode(i, QHeaderView.Stretch)
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        table.cellDoubleClicked.connect(self._on_cell_double_click)
        tab_layout.addWidget(table)

        # 按钮栏
        btn_row = QHBoxLayout()
        for label, fn in buttons_cfg:
            b = QPushButton(label)
            b.setStyleSheet(
                "font-family:'Microsoft YaHei';font-size:13px;padding:5px 10px;"
            )
            b.clicked.connect(fn)
            btn_row.addWidget(b)
        tab_layout.addLayout(btn_row)

        return {"widget": tab_widget, "table": table, "wh_type": wh_type,
                "col_defs": col_defs}

    def _current_tab(self):
        """返回当前激活 Tab 的信息 dict"""
        idx = self.tabs.currentIndex()
        if idx == 0:
            return self.reagent_tab
        elif idx == 1:
            return self.protein_tab
        else:
            return self.cell_tab

    # ── 数据刷新 ──────────────────────────────────────
    def _refresh_all(self):
        self._refresh_tab(self.reagent_tab)
        self._refresh_tab(self.protein_tab)
        self._refresh_tab(self.cell_tab)

    def _refresh_tab(self, tab_info):
        keyword = self.search_input.text().strip()
        wh_type = tab_info["wh_type"]
        table   = tab_info["table"]
        col_defs = tab_info["col_defs"]

        data = ReagentManager.search(keyword, wh_type) if keyword else \
               ReagentManager.get_all(wh_type)
        today = datetime.now()
        table.setRowCount(0)
        try:
            table.setRowCount(len(data))
            for i, r in enumerate(data):
                try:
                    self._fill_row(table, i, r, today, col_defs)
                except Exception as e:
                    item = QTableWidgetItem(f"行错误: {e}")
                    item.setTextAlignment(Qt.AlignCenter)
                    table.setItem(i, 0, item)
        except Exception as e:
            table.setRowCount(1)
            item = QTableWidgetItem(f"读取数据失败：{e}")
            item.setTextAlignment(Qt.AlignCenter)
            table.setItem(0, 0, item)

    def _fill_row(self, table, i, r, today, col_defs):
        expire_str = r.get("expire", "2099-12-31")
        try:
            exp = datetime.strptime(expire_str, "%Y-%m-%d")
            diff = (exp - today).days
        except (ValueError, TypeError):
            exp = None
            diff = 9999

        stock = r.get("stock", 0)
        is_freeze = r.get("is_freeze", False)
        loc_code = r.get("location", "")
        loc_display = self._format_location(loc_code)

        if is_freeze:
            status_txt = "❄️ 冻结"
        elif stock == 0 and r.get("warehouse_type") != WAREHOUSE_PROTEIN:
            status_txt = "⛔ 缺货"
        elif exp is not None and exp < today:
            status_txt = "💀 过期"
        elif diff <= 30:
            status_txt = "⚠️ 近效"
        else:
            status_txt = "✅ 正常"

        # 蛋白库：用体积判断是否缺货
        if r.get("warehouse_type") == WAREHOUSE_PROTEIN:
            vol, _ = parse_volume_str(r.get("sample_volume", ""))
            if vol <= 0 and not is_freeze:
                status_txt = "⛔ 缺货"
            elif status_txt == "⛔ 缺货":
                status_txt = "✅ 正常"
        # 试剂库按体积模式：也用体积判断
        elif r.get("warehouse_type") == WAREHOUSE_REAGENT and r.get("stock_mode") == "volume":
            vol, _ = parse_volume_str(r.get("sample_volume", ""))
            if vol <= 0 and not is_freeze:
                status_txt = "⛔ 缺货"
            elif status_txt == "⛔ 缺货":
                status_txt = "✅ 正常"

        bg_color = None
        if is_freeze:
            bg_color = QColor(200, 220, 255)
        elif (stock == 0 and r.get("warehouse_type") != WAREHOUSE_PROTEIN
              and not (r.get("warehouse_type") == WAREHOUSE_REAGENT
                       and r.get("stock_mode") == "volume")) or \
             (exp is not None and exp < today):
            bg_color = QColor(255, 180, 180)
        elif diff <= 30:
            bg_color = QColor(255, 240, 150)

        for col, (field, _, _) in enumerate(col_defs):
            if field == "status":
                v = status_txt
            elif field == "location":
                v = loc_display
            elif field == "stock":
                v = str(stock)
            elif field == "sample_volume":
                v = r.get("sample_volume", "")
            elif field == "stock_mode":
                mode = r.get("stock_mode", "quantity")
                v = "📏 体积" if mode == "volume" else "🔢 数量"
            else:
                v = r.get(field, "")

            item = QTableWidgetItem(str(v))
            item.setTextAlignment(Qt.AlignCenter)
            if bg_color:
                item.setBackground(bg_color)
            # 名称列保存原始 name 用于后续操作
            if field == "name":
                item.setData(Qt.UserRole, r.get("name", ""))
            table.setItem(i, col, item)

    @staticmethod
    def _format_location(loc_code):
        if not loc_code:
            return "未指定"
        info = LOCATION_MAP.get(loc_code)
        if info:
            return f"{loc_code} {info[0]}-{info[1]}"
        return loc_code

    # ── 检索 ──────────────────────────────────────────
    def _on_search(self, text):
        self._refresh_all()

    # ── 双击单元格 ────────────────────────────────────
    def _on_cell_double_click(self, row, col):
        self._detail()

    # ── 获取选中行名称 ────────────────────────────────
    def _get_selected_name(self):
        tab = self._current_tab()
        row = tab["table"].currentRow()
        if row < 0:
            QMessageBox.information(self, "提示", "请先选中一行！")
            return None
        item = tab["table"].item(row, 0)
        if not item:
            return None
        # 优先用 UserRole 中保存的原始名称
        name = item.data(Qt.UserRole)
        if name:
            return name
        return item.text()

    # ── 试剂库：领用 ──────────────────────────────────
    def _use_reagent(self):
        name = self._get_selected_name()
        if not name:
            return
        # 检查试剂的出入库模式
        raw_data = ReagentManager.get_all()
        rec = next((r for r in raw_data if r["name"] == name), None)
        if not rec:
            QMessageBox.warning(self, "错误", "未找到该试剂！")
            return

        if rec.get("stock_mode") == "volume":
            # 按体积领用（类似蛋白库）
            sv = rec.get("sample_volume", "")
            total_vol, tube_count = parse_volume_str(sv)
            vol_str = f"当前体积：{sv}（共{total_vol}mL，{tube_count}管）"
            amount, ok = QInputDialog.getDouble(
                self, "领用（按体积）",
                f"《{name}》\n{vol_str}\n\n领用体积(mL)：",
                1.0, 0.01, total_vol, 2)
            if not ok:
                return
            ok_result, days, msg = ReagentManager.use_reagent(name, amount)
            if ok_result:
                QMessageBox.information(self, "领用成功",
                    f"《{name}》体积扣减成功！\n效期剩余 {days} 天，省着点用。")
            else:
                QMessageBox.warning(self, "领用失败", f"《{name}》{msg}，无法领用！")
        else:
            # 按数量领用（原方式）
            amount, ok = QInputDialog.getInt(self, "领用", f"《{name}》领用数量：", 1, 1)
            if not ok:
                return
            ok_result, days, msg = ReagentManager.use_reagent(name, amount)
            if ok_result:
                QMessageBox.information(self, "领用成功",
                    f"《{name}》库存扣减成功！\n效期剩余 {days} 天，省着点用。")
            else:
                QMessageBox.warning(self, "领用失败", f"《{name}》{msg}，无法领用！")
        self._refresh_all()

    # ── 蛋白库：按体积领用 ────────────────────────────
    def _use_protein(self):
        name = self._get_selected_name()
        if not name:
            return
        # 先显示当前体积信息
        rec = next((r for r in ReagentManager.get_all(WAREHOUSE_PROTEIN)
                    if r["name"] == name), None)
        if not rec:
            QMessageBox.warning(self, "错误", "未找到该蛋白！")
            return
        sv = rec.get("sample_volume", "")
        total_vol, tube_count = parse_volume_str(sv)

        # 解析各管体积，供用户选择
        parts_str = sv.replace(" ", "")
        # 乘法格式展开
        import re
        m = re.match(r'^([\d.]+)\s*\*\s*(\d+)$', parts_str)
        tube_list = []
        if m:
            unit = float(m.group(1))
            cnt = int(m.group(2))
            tube_list = [unit] * cnt
        else:
            for p in parts_str.split('+'):
                p = p.strip()
                if p:
                    try:
                        tube_list.append(float(p))
                    except ValueError:
                        pass

        # 构建提示信息
        if tube_list:
            tube_info = "、".join([f"{v}mL" for v in tube_list])
            vol_str = f"当前体积：{sv}\n各管：{tube_info}\n共{total_vol}mL，{len(tube_list)}管"
        else:
            vol_str = f"当前体积：{sv}（共{total_vol}mL，{tube_count}管）"

        amount, ok = QInputDialog.getDouble(
            self, "领用蛋白",
            f"《{name}》\n{vol_str}\n\n领用体积(mL)：\n"
            f"（精确匹配某管体积将移除整管，如出库{tube_list[0] if tube_list else ''}mL将移除该管）",
            1.0, 0.01, total_vol, 2)
        if not ok:
            return
        ok_result, days, msg = ReagentManager.use_reagent(name, amount)
        if ok_result:
            # 显示出库后的剩余体积
            new_rec = next((r for r in ReagentManager.get_all(WAREHOUSE_PROTEIN)
                           if r["name"] == name), None)
            new_sv = new_rec.get("sample_volume", "") if new_rec else ""
            QMessageBox.information(self, "领用成功",
                f"《{name}》体积扣减成功！\n"
                f"出库：{amount}mL\n"
                f"剩余体积：{new_sv}\n"
                f"效期剩余 {days} 天，省着点用。")
        else:
            QMessageBox.warning(self, "领用失败", f"《{name}》{msg}！")
        self._refresh_all()

    # ── 细胞库：领用 ──────────────────────────────────
    def _use_cell(self):
        name = self._get_selected_name()
        if not name:
            return
        amount, ok = QInputDialog.getInt(self, "领用细胞", f"《{name}》领用管数：", 1, 1)
        if not ok:
            return
        ok_result, days, msg = ReagentManager.use_reagent(name, amount)
        if ok_result:
            QMessageBox.information(self, "领用成功",
                f"《{name}》扣减成功！\n效期剩余 {days} 天。")
        else:
            QMessageBox.warning(self, "领用失败", f"《{name}》{msg}！")
        self._refresh_all()

    # ── 试剂库：入库 ──────────────────────────────────
    def _add_reagent(self):
        dlg = AddReagentDialog(self)
        if dlg.exec_() == QDialog.Accepted:
            d = dlg.get_data()
            ReagentManager.add_reagent(
                d["name"], d["batch"], d["expire"], d["stock"],
                d["price"], d["threshold"], d["location"],
                warehouse_type=WAREHOUSE_REAGENT,
                specification=d.get("specification", ""),
                catalog_no=d.get("catalog_no", ""),
                manufacturer=d.get("manufacturer", ""),
                notes=d.get("notes", ""),
                stock_mode=d.get("stock_mode", "quantity"),
                sample_volume=d.get("sample_volume", ""))
            self._refresh_all()
            if d.get("stock_mode") == "volume":
                QMessageBox.information(self, "入库成功",
                    f"《{d['name']}》已入库！体积：{d.get('sample_volume', '')}mL")
            else:
                QMessageBox.information(self, "入库成功",
                    f"《{d['name']}》已入库 {d['stock']} 支！")

    # ── 蛋白库：入库 ──────────────────────────────────
    def _add_protein(self):
        dlg = AddProteinDialog(self)
        if dlg.exec_() == QDialog.Accepted:
            d = dlg.get_data()
            ReagentManager.add_reagent(
                d["name"], d["batch"], d["expire"], 0,
                0.0, 1, d["location"],
                warehouse_type=WAREHOUSE_PROTEIN,
                record_book_no=d["record_book_no"],
                project_no=d["project_no"],
                volume=d["volume"],
                molecular_weight=d["molecular_weight"],
                isoelectric_point=d["isoelectric_point"],
                extinction_coeff=d["extinction_coeff"],
                concentration=d["concentration"],
                sample_volume=d["sample_volume"],
                total_amount=d["total_amount"],
                titer=d["titer"],
                purity=d["purity"],
                buffer_type=d["buffer_type"],
                shipping_info=d["shipping_info"])
            self._refresh_all()
            QMessageBox.information(self, "入库成功",
                f"《{d['name']}》已入库！体积：{d['sample_volume']}mL")

    # ── 细胞库：入库 ──────────────────────────────────
    def _add_cell(self):
        dlg = AddCellDialog(self)
        if dlg.exec_() == QDialog.Accepted:
            d = dlg.get_data()
            ReagentManager.add_reagent(
                d["name"], d["batch"], d["expire"], d["stock"],
                0.0, 1, d["location"],
                warehouse_type=WAREHOUSE_CELL,
                cell_line=d["cell_line"],
                passage_no=d["passage_no"],
                culture_medium=d["culture_medium"],
                cryo_vial_count=d["stock"],
                viability=d["viability"],
                mycoplasma=d["mycoplasma"],
                specification=d.get("specification", ""),
                catalog_no=d.get("catalog_no", ""),
                manufacturer=d.get("manufacturer", ""),
                notes=d.get("notes", ""))
            self._refresh_all()
            QMessageBox.information(self, "入库成功",
                f"《{d['name']}》已入库 {d['stock']} 管！")

    # ── 删除 ──────────────────────────────────────────
    def _delete(self):
        name = self._get_selected_name()
        if not name:
            return
        reply = QMessageBox.question(
            self, "确认删除",
            f"确定要删除《{name}》吗？\n\n⚠️ 删除后不可恢复！",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            if ReagentManager.delete_reagent(name):
                self._refresh_all()
                QMessageBox.information(self, "删除成功", f"《{name}》已移除。")
            else:
                QMessageBox.warning(self, "删除失败", f"未找到《{name}》。")

    # ── 编辑 ──────────────────────────────────────────
    def _edit(self):
        name = self._get_selected_name()
        if not name:
            return
        raw_data = ReagentManager.get_all()
        rec = next((r for r in raw_data if r["name"] == name), None)
        if not rec:
            QMessageBox.warning(self, "错误", "找不到该数据！")
            return

        dlg = EditReagentDialog(rec, self)
        if dlg.exec_() == QDialog.Accepted:
            new_data = dlg.get_data()
            ok, msg = ReagentManager.update_reagent_full(name, new_data)
            if ok:
                self._refresh_all()
                QMessageBox.information(self, "修改成功",
                    f"《{new_data['name']}》信息已更新！")
            else:
                QMessageBox.warning(self, "修改失败", msg)

    # ── 详情 ──────────────────────────────────────────
    def _detail(self):
        name = self._get_selected_name()
        if not name:
            return
        raw_data = ReagentManager.get_all()
        rec = next((r for r in raw_data if r["name"] == name), None)
        if not rec:
            QMessageBox.warning(self, "错误", "找不到该数据！")
            return

        wh_type = rec.get("warehouse_type", WAREHOUSE_REAGENT)
        wh_label = WAREHOUSE_LABELS.get(wh_type, "未知")

        lines = [f"【{wh_label}】", f"名称：{rec.get('name', '')}"]

        if wh_type == WAREHOUSE_PROTEIN:
            # 蛋白库详情 — 按标黄字段展示
            field_order = [
                ("record_book_no",   "记录本编号"),
                ("project_no",       "项目号"),
                ("name",             "蛋白名"),
                ("volume",           "表达体积"),
                ("molecular_weight", "分子量(kD)"),
                ("isoelectric_point","等电点"),
                ("extinction_coeff", "消光系数"),
                ("concentration",    "蛋白浓度(mg/mL)"),
                ("sample_volume",    "体积(mL)"),
                ("total_amount",     "总量(mg)"),
                ("batch",            "批号"),
                ("titer",            "Titer(mg/L)"),
                ("purity",           "纯度"),
                ("buffer_type",      "Buffer"),
                ("shipping_info",    "Shipping"),
                ("expire",           "效期"),
                ("location",         "位置"),
                ("is_freeze",        "状态"),
            ]
        elif wh_type == WAREHOUSE_CELL:
            field_order = [
                ("name",           "名称"),
                ("cell_line",      "细胞系"),
                ("passage_no",     "代次"),
                ("culture_medium", "培养基"),
                ("stock",          "库存(管)"),
                ("viability",      "存活率"),
                ("mycoplasma",     "支原体检测"),
                ("specification",  "规格"),
                ("catalog_no",     "货号"),
                ("manufacturer",   "厂家"),
                ("notes",          "备注"),
                ("expire",         "效期"),
                ("location",       "位置"),
                ("is_freeze",      "状态"),
            ]
        else:
            field_order = [
                ("name",      "名称"),
                ("batch",     "批号"),
                ("expire",    "效期"),
                ("stock",     "库存"),
                ("threshold", "阈值"),
                ("stock_mode","出入库模式"),
                ("sample_volume", "体积(mL)"),
                ("price",     "单价"),
                ("specification", "规格"),
                ("catalog_no",    "货号"),
                ("manufacturer",  "厂家"),
                ("notes",         "备注"),
                ("putaway_date","入库日期"),
                ("location",  "位置"),
                ("is_freeze", "状态"),
            ]

        for field, label in field_order:
            v = rec.get(field, "")
            if field == "location":
                v = self._format_location(v) if v else "未指定"
            elif field == "is_freeze":
                v = "❄️ 已冻结" if v else "正常"
            elif field == "stock":
                v = str(v)
            elif field == "stock_mode":
                v = "按体积" if v == "volume" else "按数量"
            if v and str(v).strip():
                lines.append(f"{label}：{v}")

        detail_text = "\n".join(lines)
        QMessageBox.information(self, f"详情 — {name}", detail_text)

    # ── 冻结/解冻 ────────────────────────────────────
    def _freeze(self):
        name = self._get_selected_name()
        if not name:
            return
        frozen = ReagentManager.toggle_freeze(name)
        self._refresh_all()
        state = "已冻结，不可领用" if frozen else "已解冻，恢复正常"
        QMessageBox.information(self, "冻结状态", f"《{name}》{state}。")

    # ── 修改位置 ──────────────────────────────────────
    def _edit_location(self):
        name = self._get_selected_name()
        if not name:
            return
        self._edit_location_for(name)

    def _edit_location_for(self, name):
        loc_codes = [""] + list(LOCATION_MAP.keys())
        loc_labels = ["不指定"] + [
            f"{k} - {v[0]} {v[1]}" for k, v in LOCATION_MAP.items()
        ]
        loc, ok = QInputDialog.getItem(
            self, "修改位置", f"《{name}》的存放位置：", loc_labels, 0, False)
        if ok:
            idx = loc_labels.index(loc) if loc in loc_labels else 0
            loc_code = loc_codes[idx] if idx < len(loc_codes) else ""
            ReagentManager.update_location(name, loc_code)
            self._refresh_all()

    # ── Excel 导入 ────────────────────────────────────
    def _import_excel(self, file_path=None):
        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(self, "缺少依赖",
                "Excel 导入需要 openpyxl 库。\n请运行：pip install openpyxl")
            return

        if not file_path:
            file_path, _ = QFileDialog.getOpenFileName(
                self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)")
            if not file_path:
                return

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            field_map = {}
            used_fields = {}

            _CONFLICT_REMAP = {
                ("体积", 4): "volume",
                ("体积", 9): "sample_volume",
                ("体积(mL)", 9): "sample_volume",
            }

            for col_idx, raw_h in enumerate(header_row):
                if raw_h is None:
                    continue
                h_str = str(raw_h).replace("\n", " ").strip()
                field = _map_header(h_str)
                if not field:
                    continue
                if field in used_fields:
                    conflict_key = (h_str, col_idx + 1)
                    if conflict_key in _CONFLICT_REMAP:
                        field = _CONFLICT_REMAP[conflict_key]
                    else:
                        continue
                field_map[col_idx] = field
                used_fields[field] = col_idx

            if not field_map:
                QMessageBox.warning(self, "无法识别",
                    "Excel 表头无法匹配到库存字段！")
                wb.close()
                return

            records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rec = {}
                for col_idx, field in field_map.items():
                    val = row[col_idx] if col_idx < len(row) else None
                    if val is not None:
                        if field in ("stock", "threshold", "cryo_vial_count"):
                            try:
                                val = int(float(_strip_unit(val)))
                            except (ValueError, TypeError):
                                val = 0
                        elif field == "price":
                            try:
                                val = float(_strip_unit(val))
                            except (ValueError, TypeError):
                                val = 0.0
                        elif field == "is_freeze":
                            val = bool(val)
                        elif field in ("concentration", "sample_volume",
                                       "total_amount", "molecular_weight",
                                       "isoelectric_point", "extinction_coeff"):
                            # 浓度/体积等数值字段：去除单位后保存为字符串
                            val = _strip_unit(val)
                        else:
                            val = str(val).strip()
                        rec[field] = val
                if rec.get("name"):
                    # 自动判断仓库类型
                    if "warehouse_type" not in rec:
                        if rec.get("molecular_weight"):
                            rec["warehouse_type"] = WAREHOUSE_PROTEIN
                        elif rec.get("cell_line"):
                            rec["warehouse_type"] = WAREHOUSE_CELL
                        else:
                            rec["warehouse_type"] = WAREHOUSE_REAGENT
                    records.append(rec)
            wb.close()

            if not records:
                QMessageBox.information(self, "空数据", "Excel 中没有可导入的数据。")
                return

            added, merged = ReagentManager.import_from_list(records)
            self._refresh_all()
            QMessageBox.information(self, "导入成功",
                f"成功导入 {added} 条新记录，合并 {merged} 条已有记录。\n"
                f"识别列：{', '.join(field_map.values())}")

        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"读取 Excel 出错：{e}")

    # ── PDF 导入 ────────────────────────────────────
    def _import_pdf(self, file_path=None):
        """打开 PDF 智能识别录入对话框"""
        try:
            import pdfplumber
        except ImportError:
            QMessageBox.warning(self, "缺少依赖",
                "PDF 导入需要 pdfplumber 库。\n请运行：pip install pdfplumber")
            return

        if not file_path:
            file_path, _ = QFileDialog.getOpenFileName(
                self, "选择 PDF 文件", "", "PDF 文件 (*.pdf)")
            if not file_path:
                return

        dlg = PDFImportDialog(file_path, self)
        if dlg.exec_() == QDialog.Accepted:
            self._refresh_all()

    # ── 拖放支持 ────────────────────────────────────
    def dragEnterEvent(self, event):
        """拖入事件：接受 Excel 和 PDF 文件"""
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls:
                file_path = urls[0].toLocalFile()
                if file_path.lower().endswith(('.xlsx', '.xls', '.pdf')):
                    event.acceptProposedAction()
                    return
        event.ignore()

    def dropEvent(self, event):
        """放下事件：导入 Excel 或 PDF 文件"""
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls:
                file_path = urls[0].toLocalFile()
                if file_path.lower().endswith('.pdf'):
                    self._import_pdf(file_path)
                    return
                elif file_path.lower().endswith(('.xlsx', '.xls')):
                    self._import_excel(file_path)
                    return
        event.ignore()

# ── 入库对话框 ────────────────────────────────────────

class AddReagentDialog(QDialog):
    """试剂入库对话框"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🧪 试剂入库")
        self.setWindowIcon(QIcon(ICON_PATH))
        self.setMinimumWidth(380)

        layout = QFormLayout(self)
        layout.setLabelAlignment(Qt.AlignRight)

        self.name_edit = QLineEdit()
        self.name_edit.setToolTip("试剂名称（必填）")
        layout.addRow("名称：", self.name_edit)

        self.batch_edit = QLineEdit()
        layout.addRow("批号：", self.batch_edit)

        self.expire_edit = QLineEdit("2099-12-31")
        self.expire_edit.setToolTip("格式：YYYY-MM-DD")
        layout.addRow("效期：", self.expire_edit)

        # 出入库模式选择
        self.mode_combo = QComboBox()
        self.mode_combo.addItems(["按数量", "按体积"])
        self.mode_combo.setCurrentIndex(0)
        self.mode_combo.currentIndexChanged.connect(self._on_mode_changed)
        layout.addRow("出入库方式：", self.mode_combo)

        # 按数量时的字段
        self.stock_spin = QSpinBox()
        self.stock_spin.setRange(0, 999999)
        self.stock_spin.setValue(1)
        layout.addRow("数量：", self.stock_spin)

        self.threshold_spin = QSpinBox()
        self.threshold_spin.setRange(0, 999999)
        self.threshold_spin.setValue(1)
        layout.addRow("阈值：", self.threshold_spin)

        # 按体积时的字段
        self.sv_edit = QLineEdit()
        self.sv_edit.setToolTip(
            "多管不同体积用+号：7.26+0.2\n"
            "同体积分装用*号：5*3")
        self.sv_edit.setVisible(False)
        self.sv_label = QtLabel("体积(mL)：")
        self.sv_label.setVisible(False)
        layout.addRow(self.sv_label, self.sv_edit)

        self.price_edit = QLineEdit("0")
        layout.addRow("单价(¥)：", self.price_edit)

        self.spec_edit = QLineEdit()
        self.spec_edit.setToolTip("如：500U/管、100mL/瓶")
        layout.addRow("规格：", self.spec_edit)

        self.catalog_edit = QLineEdit()
        layout.addRow("货号：", self.catalog_edit)

        self.mfg_edit = QLineEdit()
        layout.addRow("厂家：", self.mfg_edit)

        self.notes_edit = QLineEdit()
        layout.addRow("备注：", self.notes_edit)

        # 存放位置
        self.loc_codes = [""] + list(LOCATION_MAP.keys())
        loc_labels = ["不指定"] + [
            f"{k} - {v[0]} {v[1]}" for k, v in LOCATION_MAP.items()
        ]
        self.loc_combo = QComboBox()
        self.loc_combo.addItems(loc_labels)
        layout.addRow("存放位置：", self.loc_combo)

        # 按钮
        btn_row = QHBoxLayout()
        ok_btn = QPushButton("入库")
        ok_btn.setStyleSheet(
            "font-family:'Microsoft YaHei';font-size:13px;padding:5px 20px;")
        ok_btn.clicked.connect(self._validate_and_accept)
        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet(
            "font-family:'Microsoft YaHei';font-size:13px;padding:5px 20px;")
        cancel_btn.clicked.connect(self.reject)
        btn_row.addStretch()
        btn_row.addWidget(ok_btn)
        btn_row.addWidget(cancel_btn)
        layout.addRow(btn_row)

    def _on_mode_changed(self, idx):
        """切换出入库模式时显示/隐藏相关字段"""
        is_volume = (idx == 1)
        self.stock_spin.setVisible(not is_volume)
        self.threshold_spin.setVisible(not is_volume)
        # 需要同时隐藏对应的label
        # QFormLayout中获取对应label
        for i in range(layout.count() if hasattr(self, 'layout') else 0):
            pass  # 用更简单的方式：控制row
        # 简化：直接控制widget的可见性
        self.sv_edit.setVisible(is_volume)
        self.sv_label.setVisible(is_volume)

    def _validate_and_accept(self):
        name = self.name_edit.text().strip()
        if not name:
            QMessageBox.warning(self, "输入有误", "试剂名称不能为空！")
            return
        if self.mode_combo.currentIndex() == 1:
            # 按体积模式，验证体积格式
            sv = self.sv_edit.text().strip()
            if sv:
                vol, cnt = parse_volume_str(sv)
                if vol <= 0:
                    QMessageBox.warning(self, "输入有误",
                        f"体积格式不正确：「{sv}」\n"
                        "支持格式：7.26+0.2（多管）或 5*3（同体积分装）")
                    return
        self.accept()

    def get_data(self):
        idx = self.loc_combo.currentIndex()
        loc_code = self.loc_codes[idx] if 0 <= idx < len(self.loc_codes) else ""
        try:
            price = float(self.price_edit.text())
        except ValueError:
            price = 0.0
        is_volume = (self.mode_combo.currentIndex() == 1)
        return {
            "name": self.name_edit.text().strip(),
            "batch": self.batch_edit.text().strip(),
            "expire": self.expire_edit.text().strip() or "2099-12-31",
            "stock": self.stock_spin.value() if not is_volume else 0,
            "threshold": self.threshold_spin.value() if not is_volume else 1,
            "price": price,
            "location": loc_code,
            "specification": self.spec_edit.text().strip(),
            "catalog_no": self.catalog_edit.text().strip(),
            "manufacturer": self.mfg_edit.text().strip(),
            "notes": self.notes_edit.text().strip(),
            "stock_mode": "volume" if is_volume else "quantity",
            "sample_volume": self.sv_edit.text().strip() if is_volume else "",
        }


class AddProteinDialog(QDialog):
    """蛋白入库对话框"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🧬 蛋白入库")
        self.setWindowIcon(QIcon(ICON_PATH))
        self.setMinimumWidth(440)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)

        form_widget = QWidget()
        layout = QFormLayout(form_widget)
        layout.setLabelAlignment(Qt.AlignRight)

        # 标黄字段（必须填写）
        self.name_edit = QLineEdit()
        self.name_edit.setToolTip("蛋白名（必填）")
        layout.addRow("*蛋白名：", self.name_edit)

        self.project_edit = QLineEdit()
        layout.addRow("*项目号：", self.project_edit)

        self.mw_edit = QLineEdit()
        layout.addRow("*分子量(kD)：", self.mw_edit)

        self.pi_edit = QLineEdit()
        layout.addRow("*等电点：", self.pi_edit)

        self.ec_edit = QLineEdit()
        layout.addRow("*消光系数：", self.ec_edit)

        self.conc_edit = QLineEdit()
        layout.addRow("*浓度(mg/mL)：", self.conc_edit)

        self.sv_edit = QLineEdit()
        self.sv_edit.setToolTip(
            "多管不同体积用+号：7.26+0.2\n"
            "同体积分装用*号：5*3")
        layout.addRow("*体积(mL)：", self.sv_edit)

        self.ta_edit = QLineEdit()
        layout.addRow("*总量(mg)：", self.ta_edit)

        self.batch_edit = QLineEdit()
        layout.addRow("*批号：", self.batch_edit)

        self.purity_edit = QLineEdit()
        layout.addRow("*纯度：", self.purity_edit)

        self.buffer_edit = QLineEdit()
        layout.addRow("*Buffer：", self.buffer_edit)

        # 非标黄字段
        layout.addRow(QtLabel("<b>── 其他信息 ──</b>"))

        self.record_edit = QLineEdit()
        layout.addRow("记录本编号：", self.record_edit)

        self.volume_edit = QLineEdit()
        layout.addRow("表达体积：", self.volume_edit)

        self.titer_edit = QLineEdit()
        layout.addRow("Titer(mg/L)：", self.titer_edit)

        self.ship_edit = QLineEdit()
        layout.addRow("Shipping：", self.ship_edit)

        self.expire_edit = QLineEdit("2099-12-31")
        layout.addRow("效期：", self.expire_edit)

        # 存放位置
        self.loc_codes = [""] + list(LOCATION_MAP.keys())
        loc_labels = ["不指定"] + [
            f"{k} - {v[0]} {v[1]}" for k, v in LOCATION_MAP.items()
        ]
        self.loc_combo = QComboBox()
        self.loc_combo.addItems(loc_labels)
        layout.addRow("存放位置：", self.loc_combo)

        scroll.setWidget(form_widget)

        main_layout = QVBoxLayout(self)
        main_layout.addWidget(scroll)

        btn_row = QHBoxLayout()
        ok_btn = QPushButton("入库")
        ok_btn.setStyleSheet(
            "font-family:'Microsoft YaHei';font-size:13px;padding:5px 20px;")
        ok_btn.clicked.connect(self._validate_and_accept)
        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet(
            "font-family:'Microsoft YaHei';font-size:13px;padding:5px 20px;")
        cancel_btn.clicked.connect(self.reject)
        btn_row.addStretch()
        btn_row.addWidget(ok_btn)
        btn_row.addWidget(cancel_btn)
        main_layout.addLayout(btn_row)

    def _validate_and_accept(self):
        name = self.name_edit.text().strip()
        if not name:
            QMessageBox.warning(self, "输入有误", "蛋白名不能为空！")
            return
        sv = self.sv_edit.text().strip()
        if sv:
            vol, cnt = parse_volume_str(sv)
            if vol <= 0:
                QMessageBox.warning(self, "输入有误",
                    f"体积格式不正确：「{sv}」\n"
                    "支持格式：7.26+0.2（多管）或 5*3（同体积分装）")
                return
        self.accept()

    def get_data(self):
        idx = self.loc_combo.currentIndex()
        loc_code = self.loc_codes[idx] if 0 <= idx < len(self.loc_codes) else ""
        return {
            "name": self.name_edit.text().strip(),
            "project_no": self.project_edit.text().strip(),
            "molecular_weight": self.mw_edit.text().strip(),
            "isoelectric_point": self.pi_edit.text().strip(),
            "extinction_coeff": self.ec_edit.text().strip(),
            "concentration": self.conc_edit.text().strip(),
            "sample_volume": self.sv_edit.text().strip(),
            "total_amount": self.ta_edit.text().strip(),
            "batch": self.batch_edit.text().strip(),
            "purity": self.purity_edit.text().strip(),
            "buffer_type": self.buffer_edit.text().strip(),
            "record_book_no": self.record_edit.text().strip(),
            "volume": self.volume_edit.text().strip(),
            "titer": self.titer_edit.text().strip(),
            "shipping_info": self.ship_edit.text().strip(),
            "expire": self.expire_edit.text().strip() or "2099-12-31",
            "location": loc_code,
        }


class AddCellDialog(QDialog):
    """细胞入库对话框"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🧫 细胞入库")
        self.setWindowIcon(QIcon(ICON_PATH))
        self.setMinimumWidth(400)

        layout = QFormLayout(self)
        layout.setLabelAlignment(Qt.AlignRight)

        self.name_edit = QLineEdit()
        self.name_edit.setToolTip("细胞名称（必填）")
        layout.addRow("名称：", self.name_edit)

        self.cell_line_edit = QLineEdit()
        layout.addRow("细胞系：", self.cell_line_edit)

        self.passage_edit = QLineEdit()
        layout.addRow("代次：", self.passage_edit)

        self.medium_edit = QLineEdit()
        layout.addRow("培养基：", self.medium_edit)

        self.stock_spin = QSpinBox()
        self.stock_spin.setRange(0, 999999)
        self.stock_spin.setValue(1)
        layout.addRow("冻存管数：", self.stock_spin)

        self.viability_edit = QLineEdit()
        self.viability_edit.setToolTip("如：95%")
        layout.addRow("存活率：", self.viability_edit)

        self.mycoplasma_edit = QComboBox()
        self.mycoplasma_edit.addItems(["未检测", "阴性", "阳性"])
        layout.addRow("支原体：", self.mycoplasma_edit)

        self.expire_edit = QLineEdit("2099-12-31")
        layout.addRow("效期：", self.expire_edit)

        self.batch_edit = QLineEdit()
        layout.addRow("批号：", self.batch_edit)

        self.spec_edit = QLineEdit()
        layout.addRow("规格：", self.spec_edit)

        self.catalog_edit = QLineEdit()
        layout.addRow("货号：", self.catalog_edit)

        self.mfg_edit = QLineEdit()
        layout.addRow("厂家：", self.mfg_edit)

        self.notes_edit = QLineEdit()
        layout.addRow("备注：", self.notes_edit)

        # 存放位置
        self.loc_codes = [""] + list(LOCATION_MAP.keys())
        loc_labels = ["不指定"] + [
            f"{k} - {v[0]} {v[1]}" for k, v in LOCATION_MAP.items()
        ]
        self.loc_combo = QComboBox()
        self.loc_combo.addItems(loc_labels)
        layout.addRow("存放位置：", self.loc_combo)

        # 按钮
        btn_row = QHBoxLayout()
        ok_btn = QPushButton("入库")
        ok_btn.setStyleSheet(
            "font-family:'Microsoft YaHei';font-size:13px;padding:5px 20px;")
        ok_btn.clicked.connect(self._validate_and_accept)
        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet(
            "font-family:'Microsoft YaHei';font-size:13px;padding:5px 20px;")
        cancel_btn.clicked.connect(self.reject)
        btn_row.addStretch()
        btn_row.addWidget(ok_btn)
        btn_row.addWidget(cancel_btn)
        layout.addRow(btn_row)

    def _validate_and_accept(self):
        name = self.name_edit.text().strip()
        if not name:
            QMessageBox.warning(self, "输入有误", "细胞名称不能为空！")
            return
        self.accept()

    def get_data(self):
        idx = self.loc_combo.currentIndex()
        loc_code = self.loc_codes[idx] if 0 <= idx < len(self.loc_codes) else ""
        return {
            "name": self.name_edit.text().strip(),
            "cell_line": self.cell_line_edit.text().strip(),
            "passage_no": self.passage_edit.text().strip(),
            "culture_medium": self.medium_edit.text().strip(),
            "stock": self.stock_spin.value(),
            "viability": self.viability_edit.text().strip(),
            "mycoplasma": self.mycoplasma_edit.currentText(),
            "expire": self.expire_edit.text().strip() or "2099-12-31",
            "batch": self.batch_edit.text().strip(),
            "location": loc_code,
            "specification": self.spec_edit.text().strip(),
            "catalog_no": self.catalog_edit.text().strip(),
            "manufacturer": self.mfg_edit.text().strip(),
            "notes": self.notes_edit.text().strip(),
        }


# ── 编辑对话框（通用）────────────────────────────────

class EditReagentDialog(QDialog):
    """编辑记录的对话框（根据仓库类型显示不同字段）"""

    def __init__(self, rec, parent=None):
        super().__init__(parent)
        self.rec = rec
        wh_type = rec.get("warehouse_type", WAREHOUSE_REAGENT)
        wh_label = WAREHOUSE_LABELS.get(wh_type, "编辑")
        self.setWindowTitle(f"✏️ 编辑 — {wh_label}")
        self.setWindowIcon(QIcon(ICON_PATH))
        self.setMinimumWidth(440)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)

        form_widget = QWidget()
        self.form = QFormLayout(form_widget)
        self.form.setLabelAlignment(Qt.AlignRight)

        self._edits = {}

        # 通用字段
        self._add_field("name", "名称", rec.get("name", ""), required=True)
        self._add_field("batch", "批号", rec.get("batch", ""))
        self._add_field("expire", "效期", rec.get("expire", "2099-12-31"))

        if wh_type == WAREHOUSE_PROTEIN:
            # 蛋白专属 — 标黄字段
            self.form.addRow(QtLabel("<b>── 蛋白专属 ──</b>"))
            self._add_field("project_no", "项目号", rec.get("project_no", ""))
            self._add_field("volume", "表达体积", rec.get("volume", ""))
            self._add_field("molecular_weight", "分子量(kD)", rec.get("molecular_weight", ""))
            self._add_field("isoelectric_point", "等电点", rec.get("isoelectric_point", ""))
            self._add_field("extinction_coeff", "消光系数", rec.get("extinction_coeff", ""))
            self._add_field("concentration", "浓度(mg/mL)", rec.get("concentration", ""))
            self._add_field("sample_volume", "体积(mL)", rec.get("sample_volume", ""),
                           tooltip="多管用+号，同体积用*号")
            self._add_field("total_amount", "总量(mg)", rec.get("total_amount", ""))
            self._add_field("titer", "Titer(mg/L)", rec.get("titer", ""))
            self._add_field("purity", "纯度", rec.get("purity", ""))
            self._add_field("buffer_type", "Buffer", rec.get("buffer_type", ""))
            self._add_field("shipping_info", "Shipping", rec.get("shipping_info", ""))
            self._add_field("record_book_no", "记录本编号", rec.get("record_book_no", ""))
            # 蛋白也支持规格/货号/厂家/备注
            self.form.addRow(QtLabel("<b>── 附加信息 ──</b>"))
            self._add_field("specification", "规格", rec.get("specification", ""))
            self._add_field("catalog_no", "货号", rec.get("catalog_no", ""))
            self._add_field("manufacturer", "厂家", rec.get("manufacturer", ""))
            self._add_field("notes", "备注", rec.get("notes", ""))
        elif wh_type == WAREHOUSE_CELL:
            # 细胞专属
            self.form.addRow(QtLabel("<b>── 细胞专属 ──</b>"))
            self._add_field("cell_line", "细胞系", rec.get("cell_line", ""))
            self._add_field("passage_no", "代次", rec.get("passage_no", ""))
            self._add_field("culture_medium", "培养基", rec.get("culture_medium", ""))
            self._add_field("viability", "存活率", rec.get("viability", ""))
            self._add_field("mycoplasma", "支原体", rec.get("mycoplasma", ""))
            self.form.addRow(QtLabel("<b>── 附加信息 ──</b>"))
            self._add_field("specification", "规格", rec.get("specification", ""))
            self._add_field("catalog_no", "货号", rec.get("catalog_no", ""))
            self._add_field("manufacturer", "厂家", rec.get("manufacturer", ""))
            self._add_field("notes", "备注", rec.get("notes", ""))
        else:
            # 试剂专属
            self._add_spin("stock", "库存", rec.get("stock", 0))
            self._add_spin("threshold", "阈值", rec.get("threshold", 1))
            self._add_field("price", "单价(¥)", str(rec.get("price", 0)))
            self.form.addRow(QtLabel("<b>── 出入库方式 ──</b>"))
            # 出入库模式选择
            self._mode_combo = QComboBox()
            self._mode_combo.addItems(["按数量", "按体积"])
            current_mode = rec.get("stock_mode", "quantity")
            self._mode_combo.setCurrentIndex(0 if current_mode == "quantity" else 1)
            self._mode_combo.currentIndexChanged.connect(self._on_edit_mode_changed)
            self.form.addRow("出入库方式：", self._mode_combo)
            self._add_field("sample_volume", "体积(mL)", rec.get("sample_volume", ""),
                           tooltip="按体积模式时填写，多管用+号，同体积用*号")
            self.form.addRow(QtLabel("<b>── 附加信息 ──</b>"))
            self._add_field("specification", "规格", rec.get("specification", ""))
            self._add_field("catalog_no", "货号", rec.get("catalog_no", ""))
            self._add_field("manufacturer", "厂家", rec.get("manufacturer", ""))
            self._add_field("notes", "备注", rec.get("notes", ""))

        # 位置
        self.loc_codes = [""] + list(LOCATION_MAP.keys())
        loc_labels = ["不指定"] + [
            f"{k} - {v[0]} {v[1]}" for k, v in LOCATION_MAP.items()
        ]
        self.loc_combo = QComboBox()
        self.loc_combo.addItems(loc_labels)
        current_loc = rec.get("location", "")
        if current_loc in self.loc_codes:
            self.loc_combo.setCurrentIndex(self.loc_codes.index(current_loc))
        self.form.addRow("存放位置：", self.loc_combo)

        scroll.setWidget(form_widget)

        main_layout = QVBoxLayout(self)
        main_layout.addWidget(scroll)

        btn_row = QHBoxLayout()
        ok_btn = QPushButton("确定")
        ok_btn.setStyleSheet(
            "font-family:'Microsoft YaHei';font-size:13px;padding:5px 20px;")
        ok_btn.clicked.connect(self._validate_and_accept)
        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet(
            "font-family:'Microsoft YaHei';font-size:13px;padding:5px 20px;")
        cancel_btn.clicked.connect(self.reject)
        btn_row.addStretch()
        btn_row.addWidget(ok_btn)
        btn_row.addWidget(cancel_btn)
        main_layout.addLayout(btn_row)

    def _add_field(self, key, label, value, required=False, tooltip=""):
        edit = QLineEdit(str(value))
        if required:
            edit.setToolTip("必填")
        if tooltip:
            edit.setToolTip(tooltip)
        self.form.addRow(f"{'*' if required else ''}{label}：", edit)
        self._edits[key] = edit

    def _add_spin(self, key, label, value):
        spin = QSpinBox()
        spin.setRange(0, 999999)
        spin.setValue(int(value))
        self.form.addRow(f"{label}：", spin)
        self._edits[key] = spin

    def _on_edit_mode_changed(self, idx):
        """编辑对话框中切换出入库模式"""
        is_volume = (idx == 1)
        # 更新 sample_volume 字段的提示
        if "sample_volume" in self._edits:
            sv_edit = self._edits["sample_volume"]
            sv_edit.setEnabled(is_volume)
            if not is_volume:
                sv_edit.setText("")

    def _validate_and_accept(self):
        name = self._edits["name"].text().strip()
        if not name:
            QMessageBox.warning(self, "输入有误", "名称不能为空！")
            return
        self.accept()

    def get_data(self):
        idx = self.loc_combo.currentIndex()
        loc_code = self.loc_codes[idx] if 0 <= idx < len(self.loc_codes) else ""
        result = {
            "warehouse_type": self.rec.get("warehouse_type", WAREHOUSE_REAGENT),
            "location": loc_code,
        }
        # 如果是试剂库，添加 stock_mode
        if self.rec.get("warehouse_type", WAREHOUSE_REAGENT) == WAREHOUSE_REAGENT:
            if hasattr(self, '_mode_combo'):
                result["stock_mode"] = "volume" if self._mode_combo.currentIndex() == 1 else "quantity"
        for key, widget in self._edits.items():
            if isinstance(widget, QSpinBox):
                result[key] = widget.value()
            else:
                result[key] = widget.text().strip()
        return result

# ── 独立导入函数（供 main_window 直接调用）────────────

def _import_excel_standalone(parent):
    """独立 Excel 导入函数，不需要打开 StockPanel"""
    try:
        import openpyxl
    except ImportError:
        QMessageBox.warning(parent, "缺少依赖",
            "Excel 导入需要 openpyxl 库。\n请运行：pip install openpyxl")
        return

    path, _ = QFileDialog.getOpenFileName(
        parent, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)")
    if not path:
        return

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        field_map = {}
        used_fields = {}

        _CONFLICT_REMAP = {
            ("体积", 4): "volume",
            ("体积", 9): "sample_volume",
            ("体积(mL)", 9): "sample_volume",
        }

        for col_idx, raw_h in enumerate(header_row):
            if raw_h is None:
                continue
            h_str = str(raw_h).replace("\n", " ").strip()
            field = _map_header(h_str)
            if not field:
                continue
            if field in used_fields:
                conflict_key = (h_str, col_idx + 1)
                if conflict_key in _CONFLICT_REMAP:
                    field = _CONFLICT_REMAP[conflict_key]
                else:
                    continue
            field_map[col_idx] = field
            used_fields[field] = col_idx

        if not field_map:
            QMessageBox.warning(parent, "无法识别",
                "Excel 表头无法匹配到库存字段！")
            wb.close()
            return

        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rec = {}
            for col_idx, field in field_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None:
                    if field in ("stock", "threshold", "cryo_vial_count"):
                        try:
                            val = int(float(_strip_unit(val)))
                        except (ValueError, TypeError):
                            val = 0
                    elif field == "price":
                        try:
                            val = float(_strip_unit(val))
                        except (ValueError, TypeError):
                            val = 0.0
                    elif field == "is_freeze":
                        val = bool(val)
                    elif field in ("concentration", "sample_volume",
                                   "total_amount", "molecular_weight",
                                   "isoelectric_point", "extinction_coeff"):
                        val = _strip_unit(val)
                    else:
                        val = str(val).strip()
                    rec[field] = val
            if rec.get("name"):
                if "warehouse_type" not in rec:
                    if rec.get("molecular_weight"):
                        rec["warehouse_type"] = WAREHOUSE_PROTEIN
                    elif rec.get("cell_line"):
                        rec["warehouse_type"] = WAREHOUSE_CELL
                    else:
                        rec["warehouse_type"] = WAREHOUSE_REAGENT
                records.append(rec)
        wb.close()

        if not records:
            QMessageBox.information(parent, "空数据", "Excel 中没有可导入的数据。")
            return

        added, merged = ReagentManager.import_from_list(records)
        QMessageBox.information(
            parent, "导入成功",
            f"成功导入 {added} 条新记录，合并 {merged} 条已有记录。\n"
            f"识别列：{', '.join(field_map.values())}")
    except Exception as e:
        QMessageBox.critical(parent, "导入失败", f"读取 Excel 出错：{e}")
