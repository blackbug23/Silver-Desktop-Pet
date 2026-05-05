# -*- coding: utf-8 -*-
"""
import_protein_excel.py
直接导入 20260430蛋白发货信息.xlsx 到蛋白库

运行：
    cd D:\learn\AI\Desktop-pet\DesktopPetSilver
    python import_protein_excel.py
"""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))

from openpyxl import load_workbook
from desktop_pet.reagents.manager import (
    ReagentManager, WAREHOUSE_PROTEIN, WAREHOUSE_REAGENT
)

EXCEL_PATH = r"D:\learn\AI\Desktop-pet\20260430蛋白发货信息.xlsx"

# Excel 列（0-indexed）→ 内部字段映射
COL_MAP = {
    0:  "record_book_no",   # 记录本编号
    1:  "project_no",       # 项目号
    2:  "name",            # 蛋白名 → 试剂名称
    3:  "volume",           # 体积（目标体积）
    4:  "molecular_weight", # 分子量
    5:  "isoelectric_point",# 等电点
    6:  "extinction_coeff", # 消光系数
    7:  "concentration",     # 蛋白浓度 mg/ml
    8:  "sample_volume",    # 体积（实际取样体积）
    9:  "total_amount",     # 总量
    10: "batch",            # 批号
    11: "titer",            # titer mg/L
    12: "purity",           # 纯度
    13: "buffer_type",      # buffer
    14: "shipping_info",   # shipping1
}


def main():
    print(f"读取 Excel：{EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    # 验证表头
    header = [str(c.value).strip().replace("\n", " ") for c in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"表头（{len(header)} 列）：{header}")

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[2]).strip() if row[2] else ""
        if not name or name == "None":
            continue
        rec = {"name": name, "warehouse_type": WAREHOUSE_PROTEIN}
        for col_idx, field in COL_MAP.items():
            if col_idx >= len(row):
                continue
            v = row[col_idx]
            if v is None:
                continue
            rec[field] = str(v).strip()
        # 设置默认值
        rec.setdefault("expire",     "2099-12-31")
        rec.setdefault("stock",      0)
        rec.setdefault("threshold",   1)
        rec.setdefault("price",      0.0)
        rec.setdefault("putaway_date", "")
        rec.setdefault("is_freeze",  False)
        rec.setdefault("location",   "")
        records.append(rec)
        print(f"  已解析：{name}")

    wb.close()
    print(f"\n共解析 {len(records)} 条蛋白记录，开始导入...")

    added, merged = ReagentManager.import_from_list(records)
    print(f"导入完成：新增 {added} 条，合并 {merged} 条")
    print(f"当前数据库共 {len(ReagentManager.get_all())} 条记录")


if __name__ == "__main__":
    main()
